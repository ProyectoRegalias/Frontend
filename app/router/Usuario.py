import os
from flask import Blueprint, request, render_template, redirect, url_for, session
from openpyxl.reader.excel import load_workbook

from app.models import cargar_usuarios, guardar_usuario, cargar_historial_usuario

usuario = Blueprint('main', __name__)


@usuario.route('/register', methods=['GET', 'POST'])
def register():
    error = None
    success = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        usuarios = cargar_usuarios()

        if username in usuarios:
            error = 'El nombre de usuario ya existe.'
        else:
            guardar_usuario(username, password)
            success = 'Usuario registrado con éxito. ¡Ahora puedes iniciar sesión!'
            return redirect(url_for('main.login'))

    return render_template('register.html', error=error, success=success)


@usuario.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        usuarios = cargar_usuarios()

        if username in usuarios and usuarios[username] == password:
            session['logged_in'] = True
            session['username'] = username
            session['user_data'] = cargar_historial_usuario(username)
            return redirect(url_for('chat_ia.chat'))
        else:
            error = 'Usuario o contraseña incorrectos.'

    return render_template('login.html', error=error)

@usuario.route('/ver_problemas')
def ver_problemas():
    try:
        # Cargar el archivo Excel usando openpyxl
        wb = load_workbook('arbol_problemas.xlsx')
        ws = wb.worksheets[0]

        # Crear la tabla HTML manualmente respetando celdas unidas
        html_table = "<table class='table table-bordered'>"

        for row in ws.iter_rows():
            html_table += "<tr>"
            for cell in row:
                # Detectar si la celda está en una celda combinada
                merged_cell = None
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        merged_cell = merged_range
                        break

                if merged_cell and cell.coordinate == merged_cell.coord.split(":")[0]:
                    # Si la celda es el inicio de una celda unida
                    min_col, min_row, max_col, max_row = merged_cell.bounds
                    rowspan = max_row - min_row + 1
                    colspan = max_col - min_col + 1
                    html_table += f"<td rowspan='{rowspan}' colspan='{colspan}'>{cell.value}</td>"
                elif merged_cell:
                    # Si la celda está dentro de un área unida, no mostrar nada
                    continue
                else:
                    # Si es una celda normal
                    html_table += f"<td>{cell.value}</td>"

            html_table += "</tr>"

        html_table += "</table>"

    except Exception as e:
        html_table = f"Error al cargar el archivo: {str(e)}"

    return render_template('ver_problemas.html', tabla_excel=html_table)


@usuario.route('/ver_objetivos')
def ver_objetivos():
    try:
        # Cargar el archivo Excel usando openpyxl
        file_path = os.path.join(os.path.dirname(__file__), 'arbol_objetivos.xlsx')
        wb = load_workbook(file_path)
        ws = wb.worksheets[0]

        # Crear la tabla HTML manualmente respetando celdas unidas
        html_table = "<table class='table table-bordered'>"

        for row in ws.iter_rows():
            html_table += "<tr>"
            for cell in row:
                # Detectar si la celda está en una celda combinada
                merged_cell = None
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        merged_cell = merged_range
                        break

                if merged_cell and cell.coordinate == merged_cell.coord.split(":")[0]:
                    # Si la celda es el inicio de una celda unida
                    min_col, min_row, max_col, max_row = merged_cell.bounds
                    rowspan = max_row - min_row + 1
                    colspan = max_col - min_col + 1
                    html_table += f"<td rowspan='{rowspan}' colspan='{colspan}'>{cell.value}</td>"
                elif merged_cell:
                    # Si la celda está dentro de un área unida, no mostrar nada
                    continue
                else:
                    # Si es una celda normal
                    html_table += f"<td>{cell.value}</td>"

            html_table += "</tr>"

        html_table += "</table>"

    except Exception as e:
        html_table = f"Error al cargar el archivo: {str(e)}"

    return render_template('ver_objetivos.html', tabla_excel=html_table)

@usuario.route('/arbolproblema')
def arbolproblema():
    return render_template('arbol_problema.html')

@usuario.route('/arbolobjetivos')
def arbolobjetivos():
    return render_template('arbol_objetivos.html')

@usuario.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('main.login'))
