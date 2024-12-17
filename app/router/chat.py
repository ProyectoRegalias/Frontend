import os
import uuid
import openpyxl
from flask import Blueprint, send_file, session, redirect, render_template, request, url_for
from app.model.model import Proyecto
from app.utils.utils import chat_
import pandas as pd


def limpiar_texto(texto):
    return texto.strip().replace('\n', '').replace('**', '').replace('*', '').capitalize()


chat_ia = Blueprint('chat_ia', __name__)


@chat_ia.route('/', methods=['GET', 'POST'])
@chat_ia.route('/chat', methods=['GET', 'POST'])
def chat():
    if not session.get('logged_in'):
        return redirect(url_for('main.login'))

    if 'problemas' not in session:
        session['problemas'] = {}

    """session.setdefault('causas_indirectas', [])
    session.setdefault('efectos_directos', [])
    session.setdefault('causas_directas', [])
    session.setdefault('efectos_indirectos', [])
    session.setdefault('fines_directos', [])
    session.setdefault('fines_indirectos', [])
    session.setdefault('objetivos_especificos', [])
    session.setdefault('medios', [])"""

    pregunta = ""

    if 'iteraciones' not in session:
        session['iteraciones'] = 0
    if 'respuesta_valida' not in session:
        session['respuesta_valida'] = False
    if 'pregunta' not in session:
        session['pregunta'] = ""

    causas_indirectas = []
    efectos_directos = []
    causas_directas = []
    efectos_indirectos = []
    fines_directos = []
    fines_indirectos = []
    objetivos_especificos = []
    medios = []

    respuesta = ""
    arbol = ""
    max_iteraciones = 2
    problema_id = str(uuid.uuid4())

    if request.method == 'POST':

        pregunta = request.form['pregunta']

        # Si el usuario cambió la pregunta, reinicia el contador
        if session['pregunta'] != pregunta:
            session['pregunta'] = pregunta
            session['iteraciones'] = 0
            session['respuesta_valida'] = False

        if not session['respuesta_valida']:
            response_definition_probem = chat_.send_message(
                f"Evalúa la formulación del siguiente problema según la "
                f"metodología MGA: '{pregunta}'. Verifica si el problema "
                f"está claramente definido, enfocado en una necesidad específica y "
                f"evita sugerir soluciones previas. Considera si el problema es "
                f"relevante y factible de resolver. Responde 'bien_formulado' si "
                f"cumple con estos criterios, o devuelve 'mal_formulado' si no cumple "
                f"con los criterios. Un problema bien formulado tiene las siguientes características:     "
                f"Claro y conciso: El problema se establece en términos claros y fáciles de entender, evitando jerga o lenguaje ambiguo."
                f"Específico: El problema está claramente definido y delimitado, evitando generalizaciones o vaguedades."
                f"Medible: El problema se puede medir o cuantificar de alguna manera, lo que permite evaluar el progreso hacia una solución."
                f"Alcanzable: El problema es realista y se puede resolver con los recursos disponibles."
                f"Relevante: El problema es importante y vale la pena resolverlo."
                f"Con base en la realidad: El problema se basa en hechos y datos, no en opiniones o suposiciones."
                f"Orientado a la acción: El problema está formulado de una manera que sugiere posibles soluciones."
                f"Un problema mal formulado tiene las siguientes características:"
                f"Vago e impreciso: El problema no está claramente definido y es difícil de entender."
                f"General: El problema es demasiado amplio y no se centra en un tema específico."
                f"No medible: No hay forma de medir o cuantificar el problema, lo que dificulta la evaluación del progreso."
                f"Irreal: El problema no es realista y no se puede resolver con los recursos disponibles."
                f"Irrelevante: El problema no es importante o no vale la pena resolverlo."
                f"Basado en opiniones: El problema se basa en opiniones o suposiciones, no en hechos."
                f"Sin orientación: El problema no sugiere posibles soluciones."
            ).text

            if response_definition_probem == 'bien_formulado':
                session['respuesta_valida'] = True
                respuesta = "El problema esta bien Formulado"
                #arbol = generarArbolProblemas(pregunta)
                preguntaValida = pregunta
                generarRespuestasProblemas(chat_, preguntaValida, session, problema_id)

            else:
                session['iteraciones'] += 1
                if session['iteraciones'] < max_iteraciones:
                    options_problem = chat_.send_message(
                        "El problema ingresado no está bien formulado. Aquí tienes algunas opciones "
                        "para reestructurarlo o hacerlo más claro:"
                    )
                    respuesta = options_problem.text
                else:
                    session['respuesta_valida'] = True
                    respuesta = f"Problema aceptado después de {session['iteraciones']} intentos, el problema con el que se trabajara será: {pregunta}"
                    preguntaValida2 = pregunta
                    generarRespuestasProblemas(chat_, preguntaValida2, session, problema_id)
                    #arbol = generarArbolProblemas(pregunta)

        # Si el problema ya fue validado, generar el árbol de objetivos
        else:
            respuesta = "El problema ya había sido validado anteriormente"
            #arbol = generarArbolProblemas(pregunta)

        if 'problemas' in session and problema_id in session['problemas']:
            problema_data = session['problemas'][problema_id]
            causas_indirectas = problema_data.get('causas_indirectas', [])
            efectos_directos = problema_data.get('efectos_directos', [])
            causas_directas = problema_data.get('causas_directas', [])
            efectos_indirectos = problema_data.get('efectos_indirectos', [])
            fines_directos = problema_data.get('fines_directos', [])
            fines_indirectos = problema_data.get('fines_indirectos', [])
            objetivos_especificos = problema_data.get('objetivos_especificos', [])
            medios = problema_data.get('medios', [])
            arbol = generarArbolProblemas(pregunta, causas_directas, causas_indirectas, efectos_directos,
                                          efectos_indirectos)
            print("arbol", arbol)
        else:
            causas_indirectas = []
            efectos_directos = []
            causas_directas = []
            efectos_indirectos = []
            fines_directos = []
            fines_indirectos = []
            objetivos_especificos = []
            medios = []
        print(causas_indirectas, "oepmcoedmcvfkmvifunviv")

    print("causas directas", causas_directas, "indirectas", causas_indirectas, "efec dir", efectos_directos, "efec in",
          efectos_indirectos)

    return render_template(
        'form.html',
        salida=respuesta,
        problema_valido=pregunta,
        causas_indirectas=causas_indirectas,
        efectos_directos=efectos_directos,
        causas_directas=causas_directas,
        efectos_indirectos=efectos_indirectos,
        fines_directos=fines_directos,
        fines_indirectos=fines_indirectos,
        objetivos_especificos=objetivos_especificos,
        medios=medios
    )

def generar_objetivo_principal(chat_, preguntaValida, session, problema_id):
    # Definir un buen prompt para generar el objetivo principal
    prompt = f"Analiza el siguiente problema y genera el objetivo principal para resolverlo. " \
             f"El objetivo principal debe ser una meta clara y alcanzable que resuelva de manera efectiva " \
             f"el problema planteado. El problema es el siguiente: '{preguntaValida}'. " \
             f"Considera que el objetivo debe ser amplio, pero al mismo tiempo medible y alcanzable. " \
             f"Evita soluciones previas y asegúrate de que el objetivo sea relevante y específico para el problema."

    # Enviar el mensaje a la IA
    respuesta = chat_.send_message(prompt).text
    session['obPrincipal'][problema_id] = prompt

    # Limpiar y retornar la respuesta (en caso de que sea necesario)
    return limpiar_texto(respuesta)

def generarRespuestasProblemas(chat_, preguntaValida, session, problema_id):
    problemas_data = {
        'causas_indirectas': [],
        'efectos_directos': [],
        'causas_directas': [],
        'efectos_indirectos': [],
        'fines_directos': [],
        'fines_indirectos': [],
        'objetivos_especificos': [],
        'medios': []
    }
    print("pr", preguntaValida)
    problemas_data['causas_indirectas'].append(limpiar_texto(chat_.send_message(
        f"Analiza el siguiente problema bien formulado: '{preguntaValida}' "
        f"y genera todas las causas indirectas relacionadas con el problema. Las causas indirectas son factores "
        f"subyacentes que contribuyen al problema, pero que no son inmediatamente visibles o evidentes. "
        f"Por favor, responde con todas las causas indirectas relevantes separadas por comas, asegurándote de evitar "
        f"términos sensibles o controvertidos y de mantener claridad y concisión en cada causa. Ademas no incluiras"
        f"comas en el texto si solo es uno lo colocar todo se seguido"
    ).text))

    problemas_data['efectos_directos'].append(limpiar_texto(chat_.send_message(
        f"Analiza el siguiente problema bien formulado: '{preguntaValida}' "
        f"y genera cuatro efectos directos relacionado. Un efecto directo es una consecuencia "
        f"observable y claramente vinculada al problema, visible de manera inmediata. "
        f"Por favor, responde con cuatro solo efecto directo claro y conciso, "
        f"asegurándote de evitar términos sensibles o controvertidos, como son mas de uno, los vas concatenando con "
        f"coma, asi: efecto1, efecto2,etc").text))

    problemas_data['causas_directas'].append(limpiar_texto(chat_.send_message(
        f"Analiza el siguiente problema bien formulado: '{preguntaValida}' "
        f"y genera una causa directa relacionada. Una causa directa es un factor evidente "
        f"que contribuye directamente al problema, siendo fácilmente identificable. "
        f"Por favor, responde con una sola causa directa clara y concisa, "
        f"asegurándote de evitar términos sensibles o controvertidos.  Ademas no incluiras"
        f"comas en el texto si solo es uno lo colocar todo se seguido").text))

    problemas_data['efectos_indirectos'].append(limpiar_texto(chat_.send_message(
        f"Analiza el siguiente problema bien formulado: '{preguntaValida}' "
        f"y genera un efecto indirecto relacionado. Un efecto indirecto es una consecuencia "
        f"subyacente o secundaria derivada del problema, que no es inmediatamente observable. "
        f"Por favor, responde con un solo efecto indirecto claro y conciso, "
        f"asegurándote de evitar términos sensibles o controvertidos.  Ademas no incluiras"
        f"comas en el texto si solo es uno lo colocar todo se seguido").text))

    problemas_data['medios'].append(limpiar_texto(chat_.send_message(
        f"Analiza lo siguiente: '{preguntaValida}' "
        f"y sugiere  3  medios posibles para abordar el problema. Los medios son recursos, estrategias o acciones"
        f"concretas que podrían ayudar a resolver o mitigar el problema. "
        f"Por favor, responde con 3 medios relevantes separados por comas, asegurándote de evitar términos "
        f"sensibles o controvertidos y de mantener claridad y concisión en cada medio. por ejemplo los concatnas asi:"
        f"medio1, medio2. Solo es un ejemplo de como debes concatenarlos"
    ).text))

    # Fines Directos
    problemas_data['fines_directos'].append(limpiar_texto(chat_.send_message(
        f"Analiza el siguiente problema bien formulado: '{preguntaValida}' "
        f"y sugiere un fin directo relacionado con el problema. Un fin directo es un objetivo inmediato "
        f"o resultado positivo que se espera alcanzar al abordar el problema. "
        f"Por favor, responde con un solo fin claro y conciso, "
        f"asegurándote de evitar términos sensibles o controvertidos.  Ademas no incluiras"
        f"comas en el texto si solo es uno lo colocar todo se seguido").text))

    # Fines Indirectos
    problemas_data['fines_indirectos'].append(limpiar_texto(chat_.send_message(
        f"Analiza el siguiente problema bien formulado: '{preguntaValida}' "
        f"y sugiere un fin indirecto relacionado con el problema. Un fin indirecto es un resultado secundario "
        f"o beneficio adicional que se podría lograr al abordar el problema. "
        f"Por favor, responde con un solo fin claro y conciso, "
        f"asegurándote de evitar términos sensibles o controvertidos.  Ademas no incluiras"
        f"comas en el texto si solo es uno lo colocar todo se seguido").text))

    # Objetivos Específicos
    problemas_data['objetivos_especificos'].append(limpiar_texto(chat_.send_message(
        f"Analiza el siguiente problema bien formulado: '{preguntaValida}' "
        f"y sugiere un objetivo específico relacionado con el problema. Un objetivo específico es una meta clara "
        f"y medible que contribuye a resolver el problema de manera directa. "
        f"Por favor, responde con un solo objetivo claro y conciso, "
        f"asegurándote de evitar términos sensibles o controvertidos. Ademas no incluiras"
        f"comas en el texto si solo es uno lo colocar todo se seguido").text))

    session['problemas'][problema_id] = problemas_data
    print("session", problemas_data['causas_directas'])
    print("causas_directas", problemas_data['causas_indirectas'])


def generarArbolProblemas(message_problem, causas_dir, causas_in, efectos_dir, efectos_ind):
    file_path = os.path.join(os.path.dirname(__file__), 'arbol_problemas.xlsx')

    # Verifica si el archivo existe
    if not os.path.exists(file_path):
        # Crea un DataFrame vacío y guarda un nuevo archivo .xlsx
        df = pd.DataFrame(columns=['Problema', 'Descripción', 'Categoría'])
        df.to_excel(file_path, index=False)
        print(f"Archivo creado en: {file_path}")

    tree_problems = openpyxl.load_workbook(file_path)
    sheet = tree_problems.active
    causes_directs = causas_dir
    causes_indirects = causas_in
    effects_directs = efectos_dir

    results = {
        "Efectos Indirectos": efectos_ind,
        "Efectos Directos": efectos_dir,
        "Problema": message_problem,
        "Causas Directas": causes_directs,
        "Causas Indirectas": causes_indirects
    }

    row = 2
    for category, values in results.items():
        for col, value in enumerate(values, start=3):
            if category == 'Problema':
                sheet.cell(row=4, column=3, value=message_problem)
            else:
                sheet.cell(row=row, column=col, value=value)
        row += 1
    tree_problems.save(file_path)
    print("acabo")
    return 'Arbol_problemas'


def generarArbolObejtivos(message_problem, fines_directos, fines_indirectos, medios, objetivos_especificos):
    file_path = os.path.join(os.path.dirname(__file__), 'arbol_objetivos.xlsx')
    tree_problems = openpyxl.load_workbook(file_path)
    sheet = tree_problems.active

    fines_directs = fines_directos
    fines_indirects = fines_indirectos
    general_objective = objetivos_especificos
    specific_objetives = objetivos_especificos
    means_objetives = medios

    list_fines_directs = fines_directs
    list_fines_indirects = fines_indirects
    list_specific_objetives = specific_objetives
    list_means_objetives = means_objetives

    results = {
        "Fines Directos": list_fines_directs,
        "Fines Inirectos": list_fines_indirects,
        "Objetivo General": general_objective,
        "Objetivos Especificos": list_specific_objetives,
        "Objetivos Medios": list_means_objetives
    }

    row = 2
    for category, values in results.items():
        for col, value in enumerate(values, start=3):
            if category == 'Objetivo General':
                sheet.cell(row=row, column=col, value=value)
            else:
                sheet.cell(row=row, column=col, value=value)
        row += 1
    tree_problems.save(file_path)

    return list_fines_directs, list_fines_indirects, list_specific_objetives, list_means_objetives


@chat_ia.route('/create_problems', methods=['GET', 'POST'])
def create_problems():
    json = request.get_json()
    proyecto = Proyecto()


@chat_ia.route('/matriz_formulacion', methods=['GET', 'POST'])
def matriz_formulacion():
    file_path = os.path.join(os.path.dirname(__file__), 'matriz_formulacion.xlsx')
    objetivo_especifico = chat_.send_message("Devolveras los objetivos especificos que haya")

    actividades = chat_.send_message(f"Dado los objetivos especificos anteriormente, ¿cuáles son las "
                                     "actividades clave de cada objetivo especifico que se deben realizar para "
                                     "alcanzarlo? Responde con una lista de actividades claras y accionables.")

    tareas = chat_.send_message(f"Dado los objetivos específicos y las actividades "
                                f"relacionadas: '{actividades.text}', ¿cuáles son las tareas específicas que se deben llevar "
                                f"a cabo para completar estas actividades? Responde con una lista breve y detallada.")

    personal_requerido = chat_.send_message(f"Para las tareas: '{tareas.text}', ¿qué perfiles de personal se necesitan "
                                            f"para ejecutarla, y cuál sería su disponibilidad recomendada? Incluye roles "
                                            f"específicos y nivel de experiencia.")

    resultados_actividad = chat_.send_message(
        f"Para las tareas: '{tareas.text}', ¿cuál debería ser el resultado o producto "
        f"concreto por cada tarea que se espera obtener al completarla? Responde "
        f"con un resultado medible y claro.")

    productos = chat_.send_message(f"Segun las actividades: '{actividades.text}', ¿cuál debería ser los productos por "
                                   f" cada actividad? teniendo en cuenta el manual sector 39 programa 3906, lo que retornes "
                                   f"que sea un resultado medible y claro.")

    medio_verificacion = chat_.send_message(f"Segun las actividades: '{actividades.text}', ¿cuál debería ser el "
                                            f" medio de verificacion del cumplimiento de la actividad? teniendo en cuenta el "
                                            f"manual sector 39 programa 3906, lo que retornes que sea un resultado "
                                            f"medible y claro.")

    list_objetivos = objetivo_especifico.text.split('\n')
    list_actividades = actividades.text.split('\n')
    list_tareas = tareas.text.split('\n')
    list_personal_requerido = personal_requerido.text.split('\n')
    list_resultados = resultados_actividad.text.split('\n')
    list_productos = productos.text.split('\n')
    list_medio_verificacion = medio_verificacion.text.split('\n')

    results = {
        "Objetivos": list_objetivos,
        "Actividades": list_actividades,
        "Tareas": list_tareas,
        "Personal Requerido": list_personal_requerido,
        "Resultados de Actividades": list_resultados,
        "productos": list_productos,
        "medio verificacion": list_medio_verificacion
    }

    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active

    encabezados = list(results.keys())
    for col, encabezado in enumerate(encabezados, start=1):
        sheet.cell(row=1, column=col, value=encabezado)

    # Determinar el número máximo de filas
    max_filas = max(len(values) for values in results.values())

    # Llenar datos
    for col, (category, values) in enumerate(results.items(), start=1):
        for row, value in enumerate(values, start=2):
            sheet.cell(row=row, column=col, value=value)

    # Guardar el archivo Excel
    wb.save(file_path)
    print(f"Archivo Excel guardado en: {file_path}")

    return send_file(
        file_path,
        as_attachment=True,
        download_name='matriz_formulacion.xlsx',
        mimetype='application/octet-stream'
    )
